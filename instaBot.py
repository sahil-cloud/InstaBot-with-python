# creating a insta bot to login and likes all the posts and even search for the user and can also send messages


from selenium import webdriver
from time import sleep
import openpyxl
from selenium.webdriver.common.keys import Keys

path = 'C:\\python\\jarvis\\chromedriver.exe'
excel = 'C:\\python\\Advance_python_projects\\instaBot\\Book1.xlsx'

# giving path to the driver
driver = webdriver.Chrome(path)

# excel for reading login credentials
workbook = openpyxl.load_workbook(excel)
sheet = workbook.active
n_rows = sheet.max_row

# getting userame and password from the workbook
user_data = sheet.cell(row=2, column=1).value
pass_data = sheet.cell(row=2, column=2).value


# opening the search results
driver.get('https://www.instagram.com/accounts/login/?hl=en')
sleep(4)

# geeting the username and the password
username = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/article/div/div[1]/div/form/div[2]/div/label/input')
password = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/article/div/div[1]/div/form/div[3]/div/label/input')

login_btn = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/article/div/div[1]/div/form/div[4]')

# clearing the input and sending the id password
username.clear()
password.clear()

username.send_keys(user_data)
password.send_keys(pass_data)
login_btn.click()

# clearing the both popup
sleep(5)
not_now_btn = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/div/div/div/button')
not_now_btn.click()

noti_btn = driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[3]/button[2]')
noti_btn.click()
sleep(3)

# likes = driver.find_elements_by_class_name('_8-yf5')
# # likes.click()
# # print(like)

# like_ = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/section/div/div[2]/div/article[4]/div[3]/section[1]/span[1]/button/div/svg')

# # search for the person to like the posts
# search = driver.find_element_by_xpath('//*[@id="react-root"]/section/nav/div[2]/div/div/div[2]/input')
# # the person you want to search
# search.send_keys('barkhasingh0308')
# sleep(2)
# find = driver.find_element_by_xpath('//*[@id="react-root"]/section/nav/div[2]/div/div/div[2]/div[3]/div[2]/div/a[1]')
# # search.send_keys(Keys.RETURN)
# find.click()
# sleep(3)

# driver.execute_script("window.scrollTo(0, 600)")
# sleep(3)

# # steps to like all posts
# post = driver.find_element_by_xpath(
#     '//*[@id="react-root"]/section/main/div/div[3]/article/div[1]/div/div[1]/div[1]/a/div[1]/div[2]')
# post.click()
# sleep(5)

# # we do 1st time by manually then in the loop
# like1 = driver.find_element_by_xpath(
#     '/html/body/div[4]/div[2]/div/article/div[3]/section[1]/span[1]/button')
# like1.click()
# sleep(3)
# nextpic1 = driver.find_element_by_xpath(
#     '/html/body/div[4]/div[1]/div/div/a')
# nextpic1.click()
# sleep(4)

# while True:
#     like = driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/article/div[3]/section[1]/span[1]/button')
#     like.click()
#     sleep(3)
#     nextpic = driver.find_element_by_xpath('/html/body/div[4]/div[1]/div/div/a[2]')
#     nextpic.click()
#     sleep(4)

#  for liking all the posts  }}


# for sending message to someone
inbox = driver.find_element_by_xpath('//*[@id="react-root"]/section/nav/div[2]/div/div/div[3]/div/div[2]/a')
inbox.click()
sleep(5)

send = driver.find_element_by_xpath('//*[@id="react-root"]/section/div/div[2]/div/div/div[2]/div/button')
send.click()
sleep(3)

# write the username of the person you want to send msg
ip = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/div[1]/div/div[2]/input')
# enter the person you want to send msg
ip.send_keys('') 
sleep(3)

person = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/div[2]/div/div')
# print(person)
person.click()

submit = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[1]/div/div[2]/div')
submit.click()
sleep(5)

text = driver.find_element_by_xpath('//*[@id="react-root"]/section/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/textarea')
text.send_keys('Aur kya krra hai')
text.send_keys(Keys.RETURN)
